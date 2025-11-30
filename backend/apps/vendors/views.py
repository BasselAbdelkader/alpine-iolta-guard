from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.views import View
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse_lazy
from django.contrib import messages
from django.db.models import Q
from django.http import JsonResponse, HttpResponse
from django.template.loader import render_to_string
from django.shortcuts import get_object_or_404
from .models import Vendor, VendorType
from .forms import VendorForm
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

class IndexView(LoginRequiredMixin, ListView):
    model = Vendor
    template_name = 'vendors/index.html'
    context_object_name = 'vendors'
    
    def get_queryset(self):
        queryset = Vendor.objects.select_related('vendor_type', 'client')
        
        # Search functionality
        search_query = self.request.GET.get('search', '')
        if search_query:
            queryset = queryset.filter(
                Q(vendor_name__icontains=search_query) |
                Q(contact_person__icontains=search_query) |
                Q(email__icontains=search_query) |
                Q(phone__icontains=search_query) |
                Q(vendor_number__icontains=search_query)
            )
        
        # Vendor type filter
        vendor_type = self.request.GET.get('vendor_type', '')
        if vendor_type:
            queryset = queryset.filter(vendor_type_id=vendor_type)
        
        # Client status filter (is this vendor also a client?)
        client_status = self.request.GET.get('client_status', '')
        if client_status == 'client_only':
            queryset = queryset.filter(client__isnull=False)
        elif client_status == 'vendor_only':
            queryset = queryset.filter(client__isnull=True)
        
        return queryset.order_by('-created_at')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['vendor_types'] = VendorType.objects.filter(is_active=True).order_by('name')
        context['search_query'] = self.request.GET.get('search', '')
        context['selected_vendor_type'] = self.request.GET.get('vendor_type', '')
        context['selected_client_status'] = self.request.GET.get('client_status', '')
        return context

class VendorDetailView(LoginRequiredMixin, DetailView):
    model = Vendor
    template_name = 'vendors/detail.html'
    context_object_name = 'vendor'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Get vendor payment history from trust account
        from ..bank_accounts.models import BankTransaction
        from django.db.models import Q
        from django.core.paginator import Paginator

        # Get all transactions for this vendor (payments TO this vendor)
        # Payment Register should only show WITHDRAWALS (money paid TO the vendor)
        # Include both vendor FK matches and payee name matches since they represent the same thing
        payment_items = BankTransaction.objects.filter(
            Q(vendor=self.object) |  # Vendor FK is set
            Q(payee__iexact=self.object.vendor_name)  # OR payee name matches vendor name
        ).filter(
            transaction_type__in=['WITHDRAWAL', 'TRANSFER_OUT']  # Only show payments TO vendor
        ).exclude(status='voided').select_related('bank_account', 'client').order_by('-transaction_date', '-created_at')

        # Date range filtering
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')

        if date_from:
            payment_items = payment_items.filter(transaction_date__gte=date_from)
        if date_to:
            payment_items = payment_items.filter(transaction_date__lte=date_to)

        # Search/filter by client name
        search_query = self.request.GET.get('search', '').strip()
        if search_query:
            payment_items = payment_items.filter(
                Q(client__first_name__icontains=search_query) |
                Q(client__last_name__icontains=search_query)
            )

        # Build vendor payment register
        # All items are already filtered to WITHDRAWAL/TRANSFER_OUT in the query above
        payment_register = []
        total_payments = 0

        for transaction in payment_items:
            total_payments += transaction.amount

            payment_register.append({
                'date': transaction.transaction_date,
                'amount': transaction.amount,
                'description': transaction.description,
                'reference': transaction.reference_number,
                'client': transaction.client,  # Which client's funds were used
                'client_name': transaction.client.full_name if transaction.client else 'Unassigned',
                'transaction': transaction,
                'running_total': total_payments
            })

        # Paginate payment register
        page_number = self.request.GET.get('page', 1)
        per_page = self.request.GET.get('per_page', 20)
        paginator = Paginator(payment_register, per_page)
        page_obj = paginator.get_page(page_number)
        
        # Add payment summary statistics - use the same filtered queryset
        # All items are already filtered to WITHDRAWAL/TRANSFER_OUT
        from django.db.models import Sum, Count
        payment_stats = payment_items.aggregate(
            total_amount=Sum('amount'),
            payment_count=Count('id')
        )

        # Get client breakdown (which clients have paid this vendor)
        # All items are already filtered to WITHDRAWAL/TRANSFER_OUT
        client_payments = {}
        for transaction in payment_items:
            if transaction.client:
                client_name = transaction.client.full_name
                if client_name not in client_payments:
                    client_payments[client_name] = {'amount': 0, 'count': 0}
                client_payments[client_name]['amount'] += transaction.amount
                client_payments[client_name]['count'] += 1
        
        context['payment_register'] = page_obj
        context['page_obj'] = page_obj
        context['total_payments'] = payment_stats['total_amount'] or 0
        context['payment_count'] = payment_stats['payment_count'] or 0
        context['client_payments'] = client_payments
        context['search_query'] = search_query

        return context

class VendorCreateView(LoginRequiredMixin, CreateView):
    model = Vendor
    form_class = VendorForm
    template_name = 'vendors/form.html'
    success_url = reverse_lazy('vendors:index')
    
    def get(self, request, *args, **kwargs):
        """Handle GET request - return form HTML for AJAX requests"""
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            form = self.form_class()
            html = render_to_string('vendors/modal_form.html', {
                'form': form,
                'form_title': 'Add New Vendor'
            }, request=request)
            return JsonResponse({'form_html': html})
        return super().get(request, *args, **kwargs)
    
    def form_valid(self, form):
        """Handle successful form submission"""
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            # AJAX request - return JSON response
            form.save()
            return JsonResponse({'success': True, 'message': 'Vendor created successfully.'})
        else:
            # Regular request - use default behavior
            messages.success(self.request, 'Vendor created successfully.')
            return super().form_valid(form)
    
    def form_invalid(self, form):
        """Handle form validation errors"""
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            # AJAX request - return form with errors
            html = render_to_string('vendors/modal_form.html', {
                'form': form,
                'form_title': 'Add New Vendor'
            }, request=self.request)
            return JsonResponse({'success': False, 'form_html': html})
        else:
            # Regular request - use default behavior
            return super().form_invalid(form)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form_title'] = 'Add New Vendor'
        return context

class VendorUpdateView(LoginRequiredMixin, UpdateView):
    model = Vendor
    form_class = VendorForm
    template_name = 'vendors/form.html'
    success_url = reverse_lazy('vendors:index')
    
    def get(self, request, *args, **kwargs):
        """Handle GET request - return form HTML for AJAX requests"""
        self.object = self.get_object()
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            form = self.form_class(instance=self.object)
            html = render_to_string('vendors/modal_form.html', {
                'form': form,
                'form_title': f'Edit Vendor: {self.object.vendor_name}'
            }, request=request)
            return JsonResponse({'form_html': html})
        return super().get(request, *args, **kwargs)
    
    def form_valid(self, form):
        """Handle successful form submission"""
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            # AJAX request - return JSON response
            form.save()
            return JsonResponse({'success': True, 'message': 'Vendor updated successfully.'})
        else:
            # Regular request - use default behavior
            messages.success(self.request, 'Vendor updated successfully.')
            return super().form_valid(form)
    
    def form_invalid(self, form):
        """Handle form validation errors"""
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            # AJAX request - return form with errors
            html = render_to_string('vendors/modal_form.html', {
                'form': form,
                'form_title': f'Edit Vendor: {self.object.vendor_name}'
            }, request=self.request)
            return JsonResponse({'success': False, 'form_html': html})
        else:
            # Regular request - use default behavior
            return super().form_invalid(form)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form_title'] = f'Edit Vendor: {self.object.vendor_name}'
        return context

class VendorDeleteView(LoginRequiredMixin, DeleteView):
    model = Vendor
    success_url = reverse_lazy('vendors:index')
    
    def delete(self, request, *args, **kwargs):
        vendor = self.get_object()
        messages.success(request, f'Vendor "{vendor.vendor_name}" deleted successfully.')
        return super().delete(request, *args, **kwargs)


from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import login_required

@require_http_methods(["POST"])
@login_required
def create_quick_vendor(request):
    """AJAX endpoint to quickly create a new vendor with minimal information"""
    vendor_name = request.POST.get('vendor_name', '').strip()

    if not vendor_name:
        return JsonResponse({
            'success': False,
            'message': 'Vendor name is required'
        })

    # Check if vendor already exists
    if Vendor.objects.filter(vendor_name__iexact=vendor_name).exists():
        existing_vendor = Vendor.objects.get(vendor_name__iexact=vendor_name)
        return JsonResponse({
            'success': True,
            'vendor_id': existing_vendor.id,
            'vendor_name': existing_vendor.vendor_name,
            'message': f'Vendor "{vendor_name}" already exists and has been selected.'
        })

    try:
        # Create new vendor with minimal required information
        vendor = Vendor.objects.create(
            vendor_name=vendor_name,
            is_active=True
        )

        return JsonResponse({
            'success': True,
            'vendor_id': vendor.id,
            'vendor_name': vendor.vendor_name,
            'message': f'Vendor "{vendor_name}" created successfully.'
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error creating vendor: {str(e)}'
        })


@require_http_methods(["GET"])
@login_required
def search_vendors(request):
    """AJAX endpoint to search vendors for autocomplete"""
    query = request.GET.get('q', '').strip()

    if len(query) < 2:
        return JsonResponse({'vendors': []})

    # Search vendors by name
    vendors = Vendor.objects.filter(
        vendor_name__icontains=query,
        is_active=True
    ).order_by('vendor_name')[:10]  # Limit to 10 results

    vendor_data = []
    for vendor in vendors:
        vendor_data.append({
            'id': vendor.id,
            'vendor_name': vendor.vendor_name
        })

    return JsonResponse({'vendors': vendor_data})


@require_http_methods(["GET", "POST"])
@login_required
def create_vendor_modal(request):
    """Enhanced vendor creation endpoint for modal context with pre-filled name"""

    if request.method == 'GET':
        # Get vendor name from payee autocomplete if provided
        vendor_name = request.GET.get('name', '').strip()

        # Create form with pre-filled vendor name if provided
        form = VendorForm(initial={'vendor_name': vendor_name, 'is_active': True})

        # If vendor name is provided (from payee context), make it readonly
        if vendor_name:
            form.fields['vendor_name'].widget.attrs.update({
                'readonly': True,
                'style': 'background-color: #e9ecef; cursor: not-allowed;',
                'title': 'Pre-filled from payee name - cannot be edited'
            })

        html = render_to_string('vendors/modal_form.html', {
            'form': form,
            'form_title': 'Add New Vendor',
            'is_payee_context': True,  # Flag to indicate this is from payee autocomplete
            'vendor_name_locked': bool(vendor_name)  # Flag to keep field locked on re-render
        }, request=request)

        return JsonResponse({'form_html': html})

    elif request.method == 'POST':
        # Handle form submission
        try:
            form = VendorForm(request.POST)

            if form.is_valid():
                vendor = form.save()
                return JsonResponse({
                    'success': True,
                    'vendor_id': vendor.id,
                    'vendor_name': vendor.vendor_name,
                    'message': f'Vendor "{vendor.vendor_name}" created successfully.'
                })
            else:
                # Get vendor name from POST data to keep it locked on re-render
                vendor_name = request.POST.get('vendor_name', '').strip()

                # If vendor name exists, make it readonly again (it was locked initially)
                if vendor_name:
                    form.fields['vendor_name'].widget.attrs.update({
                        'readonly': True,
                        'style': 'background-color: #e9ecef; cursor: not-allowed;',
                        'title': 'Pre-filled from payee name - cannot be edited'
                    })

                # Return form with errors
                html = render_to_string('vendors/modal_form.html', {
                    'form': form,
                    'form_title': 'Add New Vendor',
                    'is_payee_context': True,
                    'vendor_name_locked': bool(vendor_name)
                }, request=request)

                return JsonResponse({
                    'success': False,
                    'form_html': html,
                    'errors': form.errors
                })
        except Exception as e:
            import traceback
            traceback.print_exc()
            return JsonResponse({
                'success': False,
                'message': f'Error creating vendor: {str(e)}'
            }, status=500)


@require_http_methods(["GET"])
@login_required
def vendor_list_api(request):
    """API endpoint to get list of all active vendors for dropdowns"""
    vendors = Vendor.objects.filter(is_active=True).order_by('vendor_name')

    vendor_data = []
    for vendor in vendors:
        vendor_data.append({
            'id': vendor.id,
            'vendor_name': vendor.vendor_name
        })

    return JsonResponse({'vendors': vendor_data})


class VendorPaymentExportView(LoginRequiredMixin, View):
    """Export vendor payment register to XLSX"""

    def get(self, request, pk):
        vendor = get_object_or_404(Vendor, pk=pk)

        from ..bank_accounts.models import BankTransaction
        from django.db.models import Q

        # Get payment items with same filters as detail view
        # Payment Register should only show WITHDRAWALS (money paid TO the vendor)
        payment_items = BankTransaction.objects.filter(
            Q(vendor=vendor) | Q(payee__iexact=vendor.vendor_name)
        ).filter(
            transaction_type__in=['WITHDRAWAL', 'TRANSFER_OUT']  # Only show payments TO vendor
        ).exclude(status='voided').select_related('client').order_by('-transaction_date', '-created_at')

        # Apply date filters
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')

        if date_from:
            payment_items = payment_items.filter(transaction_date__gte=date_from)
        if date_to:
            payment_items = payment_items.filter(transaction_date__lte=date_to)

        # Apply search filter by client name
        search_query = request.GET.get('search', '').strip()
        if search_query:
            payment_items = payment_items.filter(
                Q(client__first_name__icontains=search_query) |
                Q(client__last_name__icontains=search_query)
            )

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Payment Register"

        # Define header style - professional black on white
        header_font = Font(bold=True, color="000000", size=11)
        header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Define accounting number format
        accounting_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'

        # Write headers
        headers = ['Date', 'Client', 'Description', 'Reference', 'Amount']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            # Add border to header
            from openpyxl.styles import Border, Side
            thin_border = Border(
                bottom=Side(style='medium', color='000000')
            )
            cell.border = thin_border

        # Write data
        total = 0
        payment_count = 0
        row_num = 2
        for transaction in payment_items:
            if transaction.transaction_type in ['WITHDRAWAL', 'TRANSFER_OUT']:
                total += float(transaction.amount)
                payment_count += 1

                # Date
                ws.cell(row=row_num, column=1, value=transaction.transaction_date.strftime('%m/%d/%Y'))

                # Client
                ws.cell(row=row_num, column=2, value=transaction.client.full_name if transaction.client else 'Unassigned')

                # Description
                ws.cell(row=row_num, column=3, value=transaction.description)

                # Reference
                ws.cell(row=row_num, column=4, value=transaction.reference_number or '')

                # Amount - with accounting format
                amount_cell = ws.cell(row=row_num, column=5, value=float(transaction.amount))
                amount_cell.number_format = accounting_format

                row_num += 1

        # Add summary row with proper accounting format
        row_num += 1  # Skip one row

        # Add border above summary
        from openpyxl.styles import Border, Side
        top_border = Border(top=Side(style='medium', color='000000'))

        # Label
        label_cell = ws.cell(row=row_num, column=1, value="Total Payments:")
        label_cell.font = Font(bold=True, size=11)

        # Total amount with accounting format
        total_amount_cell = ws.cell(row=row_num, column=5, value=total)
        total_amount_cell.font = Font(bold=True, size=11)
        total_amount_cell.number_format = accounting_format
        total_amount_cell.border = top_border

        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 12  # Date
        ws.column_dimensions['B'].width = 20  # Client
        ws.column_dimensions['C'].width = 35  # Description
        ws.column_dimensions['D'].width = 15  # Reference
        ws.column_dimensions['E'].width = 18  # Amount (wider since it's the last column)

        # Create HTTP response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{vendor.vendor_name}_payments.xlsx"'

        # Save workbook to response
        wb.save(response)

        return response
