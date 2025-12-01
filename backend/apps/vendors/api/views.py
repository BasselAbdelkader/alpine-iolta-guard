from rest_framework import viewsets, filters
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny
from django_filters.rest_framework import DjangoFilterBackend
from django.db.models import Q, Sum, Count

from ..models import Vendor, VendorType
from .serializers import VendorSerializer, VendorListSerializer, VendorTypeSerializer
from apps.api.permissions import IsTrustAccountUser
from apps.api.pagination import StandardResultsSetPagination


class VendorTypeViewSet(viewsets.ModelViewSet):
    """ViewSet for VendorType CRUD operations"""
    queryset = VendorType.objects.all()
    serializer_class = VendorTypeSerializer
    permission_classes = [IsAuthenticated]  # Require authentication
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['name', 'description']
    ordering_fields = ['name', 'created_at']
    ordering = ['name']


class VendorViewSet(viewsets.ModelViewSet):
    """ViewSet for Vendor CRUD operations with payment tracking"""
    queryset = Vendor.objects.all()
    serializer_class = VendorSerializer
    permission_classes = [IsAuthenticated]  # Require authentication
    pagination_class = StandardResultsSetPagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    
    search_fields = ['vendor_name', 'contact_person', 'email', 'vendor_number']
    filterset_fields = ['vendor_type', 'is_active', 'client']
    ordering_fields = ['vendor_name', 'created_at']
    ordering = ['vendor_name']
    
    def get_queryset(self):
        from django.db.models import Count, Sum, Q, Value, DecimalField
        from django.db.models.functions import Coalesce
        
        return Vendor.objects.select_related('vendor_type', 'client').annotate(
            annotated_payment_count=Count('bank_transactions', filter=Q(bank_transactions__transaction_type__in=['WITHDRAWAL', 'TRANSFER_OUT']) & ~Q(bank_transactions__status='voided'), distinct=True),
            annotated_total_paid=Coalesce(
                Sum('bank_transactions__amount', filter=Q(bank_transactions__transaction_type__in=['WITHDRAWAL', 'TRANSFER_OUT']) & ~Q(bank_transactions__status='voided')),
                Value(0, output_field=DecimalField(max_digits=15, decimal_places=2))
            )
        ).all()
    
    def get_serializer_class(self):
        if self.action == 'list':
            return VendorListSerializer
        return VendorSerializer
    
    def list(self, request, *args, **kwargs):
        """Enhanced list view with vendor summary"""
        response = super().list(request, *args, **kwargs)
        
        if hasattr(response, 'data') and 'results' in response.data:
            vendors = Vendor.objects.all()
            active_vendors = vendors.filter(is_active=True).count()
            client_vendors = vendors.filter(client__isnull=False).count()
            
            response.data['summary'] = {
                'total_vendors': vendors.count(),
                'active_vendors': active_vendors,
                'inactive_vendors': vendors.count() - active_vendors,
                'client_vendors': client_vendors,
                'regular_vendors': vendors.count() - client_vendors
            }
        
        return response
    
    @action(detail=False, methods=['get'])
    def search(self, request):
        """Enhanced search endpoint for vendors"""
        query = request.query_params.get('q', '')
        limit = int(request.query_params.get('limit', 50))
        
        if len(query) < 2:
            return Response({
                'vendors': [],
                'count': 0,
                'query': query,
                'message': 'Search query must be at least 2 characters'
            })
        
        vendors = Vendor.objects.select_related('vendor_type', 'client').filter(
            Q(vendor_name__icontains=query) |
            Q(contact_person__icontains=query) |
            Q(email__icontains=query) |
            Q(vendor_number__icontains=query)
        ).order_by('vendor_name')[:limit]
        
        serializer = VendorListSerializer(vendors, many=True)
        return Response({
            'vendors': serializer.data,
            'count': vendors.count(),
            'query': query,
            'limit': limit
        })
    
    @action(detail=True, methods=['get'])
    def payments(self, request, pk=None):
        """Get payment history for a specific vendor with client breakdown"""
        vendor = self.get_object()

        from apps.bank_accounts.models import BankTransaction
        from apps.clients.models import Client

        # Get date filters
        date_from = request.query_params.get('date_from')
        date_to = request.query_params.get('date_to')

        # Filter payments
        payments_qs = BankTransaction.objects.filter(
            vendor=vendor, transaction_type__in=['WITHDRAWAL', 'TRANSFER_OUT']
        ).exclude(status='voided').select_related('case', 'client').order_by('-transaction_date')

        if date_from:
            payments_qs = payments_qs.filter(transaction_date__gte=date_from)
        if date_to:
            payments_qs = payments_qs.filter(transaction_date__lte=date_to)

        # Build payment register with running total
        payment_data = []
        running_total = 0
        client_breakdown = {}

        for payment in payments_qs:
            running_total += payment.amount

            client_name = payment.client.full_name if payment.client else 'Unknown'

            # Track by client
            if client_name not in client_breakdown:
                client_breakdown[client_name] = {
                    'amount': 0,
                    'count': 0
                }
            client_breakdown[client_name]['amount'] += float(payment.amount)
            client_breakdown[client_name]['count'] += 1

            payment_data.append({
                'id': payment.id,
                'transaction_number': payment.transaction_number,
                'date': payment.transaction_date.strftime('%m/%d/%Y'),
                'amount': str(payment.amount),
                'running_total': str(running_total),
                'description': payment.description,
                'reference': payment.reference_number,
                'client_name': client_name,
                'client_id': payment.client.id if payment.client else None,
                'case_title': payment.case.case_title if payment.case else None,
                'status': payment.status
            })

        # Calculate percentages
        total = sum(c['amount'] for c in client_breakdown.values())
        client_breakdown_list = []
        for client_name, data in sorted(client_breakdown.items(), key=lambda x: x[1]['amount'], reverse=True):
            percentage = (data['amount'] / total * 100) if total > 0 else 0
            client_breakdown_list.append({
                'client_name': client_name,
                'amount': data['amount'],
                'count': data['count'],
                'percentage': round(percentage, 1)
            })

        return Response({
            'vendor': {
                'id': vendor.id,
                'vendor_name': vendor.vendor_name,
                'vendor_number': vendor.vendor_number,
                'contact_person': vendor.contact_person,
                'email': vendor.email,
                'phone': vendor.phone,
                'address': vendor.address,
                'city': vendor.city,
                'state': vendor.state,
                'zip_code': vendor.zip_code,
                'is_active': vendor.is_active,
                'created_at': vendor.created_at.strftime('%m/%d/%Y'),
                'updated_at': vendor.updated_at.strftime('%m/%d/%Y')
            },
            'total_payments': str(running_total),
            'payment_count': len(payment_data),
            'payments': payment_data,
            'client_breakdown': client_breakdown_list
        })

    @action(detail=True, methods=['get'], url_path='payments/export')
    def export_payments(self, request, pk=None):
        """Export vendor payment register to XLSX"""
        from django.http import HttpResponse
        import io
        from datetime import datetime
        import traceback
        from apps.vendors.models import Vendor
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment

        try:
            # Get vendor directly to bypass filter backends
            vendor = Vendor.objects.get(pk=pk)

            # Get date filters (works with both DRF Request and Django WSGIRequest)
            params = getattr(request, 'query_params', request.GET)
            date_from = params.get('date_from')
            date_to = params.get('date_to')

            # Get payments using same logic as payments endpoint
            from apps.bank_accounts.models import BankTransaction
            payments_qs = BankTransaction.objects.filter(
                vendor=vendor, transaction_type__in=['WITHDRAWAL', 'TRANSFER_OUT']
            ).exclude(status='voided').select_related('case', 'client').order_by('transaction_date')

            if date_from:
                payments_qs = payments_qs.filter(transaction_date__gte=date_from)
            if date_to:
                payments_qs = payments_qs.filter(transaction_date__lte=date_to)

            # Create Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Payment Register"

            # Set column widths (matching sample format)
            ws.column_dimensions['A'].width = 12  # Date
            ws.column_dimensions['B'].width = 20  # Client
            ws.column_dimensions['C'].width = 35  # Description
            ws.column_dimensions['D'].width = 15  # Reference
            ws.column_dimensions['E'].width = 18  # Amount

            # Header row with formatting
            headers = ['Date', 'Client', 'Description', 'Reference', 'Amount']
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')

            # Add payment data
            total_amount = 0
            row_num = 2
            for payment in payments_qs:
                total_amount += float(payment.amount)
                client_name = payment.client.full_name if payment.client else 'Unknown'

                ws.cell(row=row_num, column=1, value=payment.transaction_date.strftime('%m/%d/%Y'))
                ws.cell(row=row_num, column=2, value=client_name)
                ws.cell(row=row_num, column=3, value=payment.description or '')
                ws.cell(row=row_num, column=4, value=payment.reference_number or '')
                ws.cell(row=row_num, column=5, value=float(payment.amount))

                row_num += 1

            # Add blank row
            row_num += 1

            # Add total row with bold formatting
            total_cell = ws.cell(row=row_num, column=1, value='Total Payments:')
            total_cell.font = Font(bold=True)

            amount_cell = ws.cell(row=row_num, column=5, value=total_amount)
            amount_cell.font = Font(bold=True)

            # Save to BytesIO
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            # Return as XLSX file
            vendor_name_safe = vendor.vendor_name.replace('/', '_').replace('\\', '_')
            filename = f"{vendor_name_safe}_payments.xlsx"
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response

        except Exception as e:
            # Log the full error
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error in export_payments: {str(e)}")
            logger.error(traceback.format_exc())
            # Return error response with full traceback in development
            return HttpResponse(
                f"Error generating export:\n\n{str(e)}\n\n{traceback.format_exc()}",
                status=500,
                content_type='text/plain'
            )
